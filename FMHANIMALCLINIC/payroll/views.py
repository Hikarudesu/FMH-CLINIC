"""
Simplified Payroll Views for FMH Animal Clinic.

Flow:
1. Dashboard - overview and quick actions
2. Generate Payslips - create payslips for a period
3. View/Edit Payslips - review and adjust individual payslips
4. Release Payroll - mark as cash released
5. Print Payslip - print for employee

Tab Navigation:
- Dashboard: Overview with stats and activity
- Vets: List of employees with payroll history
- Requests: Payroll generation requests/history
- Payroll Logs: Transaction records of salary disbursements
"""
import logging
from datetime import date, timedelta
from decimal import Decimal, InvalidOperation
import json

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Sum, Count
from django.utils import timezone
from django.http import Http404

from accounts.models import User
from accounts.decorators import module_permission_required, special_permission_required
from employees.models import StaffMember
from .models import PayrollPeriod, Payslip

logger = logging.getLogger('fmh')


# ═══════════════════════════════════════════════════════════════════
#                       HELPER FUNCTIONS
# ═══════════════════════════════════════════════════════════════════

def safe_decimal(value, default=Decimal('0')):
    """Safely convert a value to Decimal, returning default on error."""
    if value is None or value == '':
        return default
    try:
        return Decimal(str(value))
    except (InvalidOperation, ValueError, TypeError):
        return default


def safe_int(value, default=0):
    """Safely convert a value to int, returning default on error."""
    if value is None or value == '':
        return default
    try:
        return int(value)
    except (ValueError, TypeError):
        return default


# ═══════════════════════════════════════════════════════════════════
#                           DASHBOARD
# ═══════════════════════════════════════════════════════════════════

@login_required
@module_permission_required('payroll', 'MANAGE')
def payroll_dashboard(request):
    """Main payroll dashboard - simple overview with optimized queries."""
    today = date.today()
    
    # Current month period
    current_period = PayrollPeriod.objects.filter(
        month=today.month, year=today.year
    ).first()
    
    # Recent periods
    recent_periods = PayrollPeriod.objects.all()[:12]
    
    # Year-to-date stats - single optimized query
    ytd_stats = {'total_gross': 0, 'total_net': 0, 'count': 0}
    try:
        ytd_result = Payslip.objects.filter(
            payroll_period__year=today.year,
            status__in=['APPROVED', 'RELEASED']
        ).aggregate(
            total_gross=Sum('gross_pay'),
            total_net=Sum('net_pay'),
            count=Count('id')
        )
        ytd_stats = {
            'total_gross': ytd_result['total_gross'] or 0,
            'total_net': ytd_result['total_net'] or 0,
            'count': ytd_result['count'] or 0
        }
    except Exception as e:
        logger.warning(f"Error fetching YTD stats: {e}")
    
    # Active employees count
    active_employees = 0
    try:
        active_employees = StaffMember.objects.filter(
            is_active=True,
            user__isnull=False,
            user__assigned_role__is_staff_role=True
        ).count()
    except Exception as e:
        logger.warning(f"Error counting active employees: {e}")
    
    # Pending payslips count
    pending_count = 0
    try:
        pending_count = Payslip.objects.filter(status='DRAFT').count()
    except Exception as e:
        logger.warning(f"Error counting pending payslips: {e}")
    
    # Last payroll cycle
    last_released = None
    try:
        last_released = PayrollPeriod.objects.filter(status='RELEASED').first()
    except Exception as e:
        logger.warning(f"Error fetching last released period: {e}")
    
    # Monthly trends - optimized single query for all 12 months
    monthly_trends = []
    try:
        # Generate month ranges
        trend_months = []
        for i in range(11, -1, -1):
            trend_date = today - timedelta(days=30*i)
            trend_months.append((trend_date.month, trend_date.year, trend_date))
        
        # Single query for all relevant periods
        released_periods = {
            (p.month, p.year): p
            for p in PayrollPeriod.objects.filter(
                status='RELEASED',
                year__gte=today.year - 1
            )
        }
        
        for month, year, trend_date in trend_months:
            period = released_periods.get((month, year))
            amount = float(period.total_net or 0) if period else 0
            monthly_trends.append({
                'month': trend_date.strftime('%b'),
                'amount': amount,
                'full_month': trend_date.strftime('%B %Y')
            })
    except Exception as e:
        logger.warning(f"Error calculating monthly trends: {e}")
        # Fallback: empty trends
        for i in range(11, -1, -1):
            trend_date = today - timedelta(days=30*i)
            monthly_trends.append({
                'month': trend_date.strftime('%b'),
                'amount': 0,
                'full_month': trend_date.strftime('%B %Y')
            })
    
    # Recent activity - single optimized query
    recent_activity = []
    try:
        recent_activity = list(Payslip.objects.filter(
            payroll_period__year=today.year
        ).select_related(
            'employee', 'payroll_period'
        ).order_by('-payroll_period__generated_at')[:5])
    except Exception as e:
        logger.warning(f"Error fetching recent activity: {e}")
    
    # Upcoming payroll (next month)
    next_month = today.month + 1 if today.month < 12 else 1
    next_year = today.year if today.month < 12 else today.year + 1
    upcoming_period = None
    try:
        upcoming_period = PayrollPeriod.objects.filter(
            month=next_month, year=next_year
        ).first()
    except Exception as e:
        logger.warning(f"Error fetching upcoming period: {e}")
    
    # Total payroll this year - can reuse ytd_stats
    total_payroll_this_year = Decimal('0')
    try:
        result = Payslip.objects.filter(
            payroll_period__year=today.year,
            status='RELEASED'
        ).aggregate(Sum('net_pay'))['net_pay__sum']
        if result:
            total_payroll_this_year = result
    except Exception as e:
        logger.warning(f"Error calculating total payroll this year: {e}")
    
    context = {
        'current_period': current_period,
        'recent_periods': recent_periods,
        'ytd_stats': ytd_stats,
        'active_employees': active_employees,
        'pending_count': pending_count,
        'last_released': last_released,
        'upcoming_period': upcoming_period,
        'monthly_trends': json.dumps(monthly_trends),
        'recent_activity': recent_activity,
        'total_payroll_this_year': total_payroll_this_year,
        'current_month': today.month,
        'current_year': today.year,
        'tab': 'dashboard',
    }
    
    return render(request, 'payroll/dashboard.html', context)


# ═══════════════════════════════════════════════════════════════════
#                      GENERATE PAYSLIPS
# ═══════════════════════════════════════════════════════════════════

@login_required
@module_permission_required('payroll', 'MANAGE')
def generate_payslips(request):
    """Generate payslips for a selected period."""
    today = date.today()
    
    # Get selected period or default to current month
    month = safe_int(request.GET.get('month', today.month), today.month)
    year = safe_int(request.GET.get('year', today.year), today.year)
    
    # Get or create period
    period, created = PayrollPeriod.objects.get_or_create(
        month=month, year=year,
        defaults={'status': 'DRAFT'}
    )
    
    # Get active employees safely
    employees = []
    employee_count = 0
    try:
        emp_ids = list(StaffMember.objects.filter(
            is_active=True,
            user__isnull=False,
            user__assigned_role__is_staff_role=True
        ).values_list('id', flat=True))
        employee_count = len(emp_ids)
        
        for emp_id in emp_ids:
            try:
                emp = StaffMember.objects.select_related('branch').get(id=emp_id)
                employees.append(emp)
            except Exception:
                continue
    except Exception:
        employees = []
        employee_count = 0
    
    # Check for existing payslips
    existing_payslips = {}
    try:
        for p in period.payslips.values_list('employee_id', 'id'):
            existing_payslips[p[0]] = p[1]
    except Exception:
        existing_payslips = {}
    
    # Month choices for dropdown
    months = [
        {'num': i, 'name': date(2000, i, 1).strftime('%B')}
        for i in range(1, 13)
    ]
    
    context = {
        'period': period,
        'employees': employees,
        'existing_payslips': existing_payslips,
        'months': months,
        'selected_month': month,
        'selected_year': year,
        'employee_count': employee_count,
    }
    
    return render(request, 'payroll/generate.html', context)


@login_required
@module_permission_required('payroll', 'MANAGE')
def generate_payslips_action(request):
    """Process the payslip generation."""
    if request.method != 'POST':
        return redirect('payroll:generate')
    
    month = safe_int(request.POST.get('month'), date.today().month)
    year = safe_int(request.POST.get('year'), date.today().year)
    
    # Get or create period
    period, _ = PayrollPeriod.objects.get_or_create(
        month=month, year=year,
        defaults={'status': 'DRAFT'}
    )
    
    # Check if already released
    if period.status == 'RELEASED':
        messages.error(request, 'This payroll period has already been released.')
        return redirect('payroll:payslips', period_id=period.id)
    
    # Get active employees safely
    created_count = 0
    updated_count = 0
    total_employees = 0
    
    try:
        emp_ids = list(StaffMember.objects.filter(
            is_active=True,
            user__isnull=False,
            user__assigned_role__is_staff_role=True
        ).values_list('id', flat=True))
        total_employees = len(emp_ids)
        
        for emp_id in emp_ids:
            try:
                emp = StaffMember.objects.get(id=emp_id)
                payslip, created = Payslip.objects.get_or_create(
                    payroll_period=period,
                    employee=emp,
                    defaults={'status': 'DRAFT'}
                )
                
                # Generate/update payslip data
                payslip.generate_from_employee()
                payslip.save()
                
                if created:
                    created_count += 1
                else:
                    updated_count += 1
            except Exception:
                # Skip employees with errors
                continue
    except Exception:
        messages.error(request, 'Error generating payslips.')
        return redirect('payroll:generate')
    
    # Update period totals and status
    period.status = 'GENERATED'
    period.generated_at = timezone.now()
    period.update_totals()
    period.save()
    
    # Log the generation
    from .models import PayrollAuditLog
    PayrollAuditLog.log(
        user=request.user,
        action_type=PayrollAuditLog.ActionType.PERIOD_GENERATED,
        description=f"Generated payslips for {period.period_display} - {created_count} new, {updated_count} updated ({total_employees} total employees)",
        payroll_period=period,
        metadata={
            'created_count': created_count,
            'updated_count': updated_count,
            'total_employees': total_employees
        }
    )
    
    messages.success(
        request, 
        f'Payslips generated! {created_count} new, {updated_count} updated.'
    )
    
    return redirect('payroll:payslips', period_id=period.id)


# ═══════════════════════════════════════════════════════════════════
#                      VIEW/LIST PAYSLIPS
# ═══════════════════════════════════════════════════════════════════

@login_required
@module_permission_required('payroll', 'MANAGE')
def payslips_list(request, period_id):
    """View all payslips for a period."""
    period = get_object_or_404(PayrollPeriod, id=period_id)
    
    # Get payslips safely, handling corrupted decimal data
    payslips = []
    try:
        # Get IDs first to avoid accessing decimal fields during query
        payslip_ids = list(period.payslips.values_list('id', flat=True))
        
        # Fetch each payslip individually with error handling
        for ps_id in payslip_ids:
            try:
                ps = Payslip.objects.select_related('employee', 'employee__branch').get(id=ps_id)
                payslips.append(ps)
            except Exception:
                # Skip payslips with corrupted data
                continue
    except Exception:
        # If we can't even get IDs, return empty list
        payslips = []
    
    context = {
        'period': period,
        'payslips': payslips,
    }
    
    return render(request, 'payroll/payslips_list.html', context)


@login_required
@module_permission_required('payroll', 'MANAGE')
def payslip_edit(request, payslip_id):
    """Edit individual payslip."""
    payslip = get_object_or_404(
        Payslip.objects.select_related('employee', 'payroll_period'),
        id=payslip_id
    )
    
    if request.method == 'POST':
        # Update allowances
        payslip.overtime_hours = safe_decimal(request.POST.get('overtime_hours', 0))
        payslip.overtime_pay = safe_decimal(request.POST.get('overtime_pay', 0))
        payslip.holiday_pay = safe_decimal(request.POST.get('holiday_pay', 0))
        payslip.bonus = safe_decimal(request.POST.get('bonus', 0))
        payslip.allowance = safe_decimal(request.POST.get('allowance', 0))
        payslip.staff_allowance = safe_decimal(request.POST.get('staff_allowance', 2000))
        
        # Update deductions (SSS/PhilHealth/PagIBIG are clinic-paid, not edited here)
        payslip.tax = safe_decimal(request.POST.get('tax', 0))
        payslip.cash_advance = safe_decimal(request.POST.get('cash_advance', 0))
        payslip.late_deduction = safe_decimal(request.POST.get('late_deduction', 0))
        payslip.absent_deduction = safe_decimal(request.POST.get('absent_deduction', 0))
        payslip.other_deductions = safe_decimal(request.POST.get('other_deductions', 0))
        
        # Update days
        payslip.days_worked = safe_int(request.POST.get('days_worked', 22), 22)
        payslip.days_absent = safe_int(request.POST.get('days_absent', 0))
        
        # Notes
        payslip.notes = request.POST.get('notes', '')
        
        # Recalculate and save
        payslip.calculate()
        payslip.save()
        
        # Update period totals
        payslip.payroll_period.update_totals()
        
        # Log the edit
        from .models import PayrollAuditLog
        PayrollAuditLog.log(
            user=request.user,
            action_type=PayrollAuditLog.ActionType.PAYSLIP_EDITED,
            description=f"Edited payslip for {payslip.employee.full_name} ({payslip.payroll_period.period_display})",
            payslip=payslip,
            payroll_period=payslip.payroll_period,
            staff_member=payslip.employee
        )
        
        messages.success(request, f'Payslip for {payslip.employee.full_name} updated.')
        return redirect('payroll:vets')
    
    context = {
        'payslip': payslip,
        'period': payslip.payroll_period,
    }
    
    return render(request, 'payroll/payslip_edit.html', context)


# ═══════════════════════════════════════════════════════════════════
#                      RELEASE PAYROLL
# ═══════════════════════════════════════════════════════════════════

@login_required
@module_permission_required('payroll', 'MANAGE')
def release_payroll(request, period_id):
    """Release payroll - mark all payslips as released (cash paid)."""
    period = get_object_or_404(PayrollPeriod, id=period_id)
    
    if request.method == 'POST':
        if period.status == 'RELEASED':
            messages.warning(request, 'This payroll has already been released.')
        else:
            from .models import PayrollAuditLog
            now = timezone.now()
            
            # Get payslip count before release
            try:
                payslip_count = period.payslips.count()
            except Exception:
                payslip_count = 0
            
            # Mark all payslips as released
            period.payslips.update(
                status='RELEASED',
                released_at=now
            )
            
            # Update period status
            period.status = 'RELEASED'
            period.released_at = now
            period.released_by = request.user
            period.save()
            
            # Log the release - with error handling for total_net
            total_net_value = Decimal('0')
            try:
                if period.total_net:
                    total_net_value = period.total_net
            except Exception:
                total_net_value = Decimal('0')
            
            PayrollAuditLog.log(
                user=request.user,
                action_type=PayrollAuditLog.ActionType.PERIOD_RELEASED,
                description=f"Released payroll for {period.period_display} - {payslip_count} payslips sent to vet accounts",
                payroll_period=period,
                metadata={
                    'employee_count': payslip_count,
                    'total_net': str(total_net_value)
                }
            )
            
            messages.success(
                request, 
                f'Payroll for {period.period_display} has been released!'
            )
        
        return redirect('payroll:payslips', period_id=period.id)
    
    # Confirmation page
    context = {
        'period': period,
    }
    return render(request, 'payroll/release_confirm.html', context)


# ═══════════════════════════════════════════════════════════════════
#                      PRINT PAYSLIP
# ═══════════════════════════════════════════════════════════════════

@login_required
@module_permission_required('payroll', 'MANAGE')
def payslip_print(request, payslip_id):
    """Print-friendly payslip view."""
    payslip = get_object_or_404(
        Payslip.objects.select_related('employee', 'employee__branch', 'payroll_period'),
        id=payslip_id
    )
    
    context = {
        'payslip': payslip,
        'print_mode': True,
    }
    
    return render(request, 'payroll/payslip_print.html', context)


# ═══════════════════════════════════════════════════════════════════
#                      QUICK ACTIONS / API
# ═══════════════════════════════════════════════════════════════════

@login_required
@module_permission_required('payroll', 'MANAGE')
def payslip_approve(request, payslip_id):
    """Approve a single payslip."""
    payslip = get_object_or_404(Payslip, id=payslip_id)
    
    if payslip.status == 'DRAFT':
        payslip.status = 'APPROVED'
        payslip.save()
        messages.success(request, f'Payslip for {payslip.employee.full_name} approved.')
    
    return redirect('payroll:payslips', period_id=payslip.payroll_period.id)


@login_required
@module_permission_required('payroll', 'MANAGE')
def approve_all_payslips(request, period_id):
    """Approve all payslips in a period."""
    period = get_object_or_404(PayrollPeriod, id=period_id)
    
    count = period.payslips.filter(status='DRAFT').update(status='APPROVED')
    messages.success(request, f'{count} payslips approved.')
    
    return redirect('payroll:payslips', period_id=period.id)


@login_required
@module_permission_required('payroll', 'MANAGE') 
def delete_period(request, period_id):
    """Delete a payroll period (only if DRAFT)."""
    period = get_object_or_404(PayrollPeriod, id=period_id)
    
    if period.status == 'RELEASED':
        messages.error(request, 'Cannot delete a released payroll period.')
        return redirect('payroll:dashboard')
    
    if request.method == 'POST':
        period_name = period.period_display
        period.delete()
        messages.success(request, f'Payroll period "{period_name}" deleted.')
        return redirect('payroll:dashboard')
    
    context = {'period': period}
    return render(request, 'payroll/delete_confirm.html', context)


# ═══════════════════════════════════════════════════════════════════
#                    TAB VIEWS - VETS (EMPLOYEES)
# ═══════════════════════════════════════════════════════════════════

@login_required
@module_permission_required('payroll', 'MANAGE')
def payroll_vets(request):
    """View all employees with their payroll information."""
    today = date.today()
    
    # Get all branches for filter dropdown
    from branches.models import Branch
    branches = Branch.objects.all().order_by('name')
    
    # Current period payslips
    current_period = PayrollPeriod.objects.filter(
        month=today.month, year=today.year
    ).first()
    
    # Get employees safely with error handling
    employee_data = []
    active_count = 0
    branch_filter = request.GET.get('branch')
    
    try:
        # Get all active employee IDs first - use values to avoid decimal fields
        if branch_filter:
            emp_ids = list(StaffMember.objects.filter(
                is_active=True,
                user__isnull=False,
                user__assigned_role__is_staff_role=True,
                branch_id=branch_filter
            ).values_list('id', flat=True))
        else:
            emp_ids = list(StaffMember.objects.filter(
                is_active=True,
                user__isnull=False,
                user__assigned_role__is_staff_role=True
            ).values_list('id', flat=True))
        
        # Try to get active count - might fail if corrupted data exists
        try:
            all_active = StaffMember.objects.filter(
                is_active=True,
                user__isnull=False,
                user__assigned_role__is_staff_role=True
            ).values_list('id')
            active_count = len(list(all_active))
        except Exception:
            active_count = 0
        
        # Build employee payroll data - handle corrupted data
        for emp_id in emp_ids:
            try:
                emp = StaffMember.objects.get(id=emp_id)
                
                # Current period payslip
                current_payslip = None
                if current_period:
                    current_payslip = Payslip.objects.filter(
                        employee=emp,
                        payroll_period=current_period
                    ).first()
                
                # Last released payslip
                last_payslip = Payslip.objects.filter(
                    employee=emp,
                    status='RELEASED'
                ).order_by('-payroll_period__released_at').first()
                
                # YTD total
                ytd_gross = Decimal('0')
                try:
                    ytd_result = Payslip.objects.filter(
                        employee=emp,
                        payroll_period__year=today.year,
                        status='RELEASED'
                    ).aggregate(Sum('gross_pay'))['gross_pay__sum']
                    if ytd_result:
                        ytd_gross = ytd_result
                except Exception:
                    ytd_gross = Decimal('0')
                
                employee_data.append({
                    'employee': emp,
                    'current_payslip': current_payslip,
                    'last_payslip': last_payslip,
                    'ytd_gross': ytd_gross,
                })
            except Exception:
                # Skip this employee if there's any error accessing their data
                continue
                
    except Exception:
        # If we can't even get the employee IDs, return empty list
        employee_data = []
    
    context = {
        'tab': 'vets',
        'employees': employee_data,
        'active_count': active_count,
        'filtered_count': len(employee_data),
        'branches': branches,
        'selected_branch': request.GET.get('branch'),
        'current_period': current_period,
    }
    
    return render(request, 'payroll/vets.html', context)


# ═══════════════════════════════════════════════════════════════════
#                    TAB VIEWS - REQUESTS
# ═══════════════════════════════════════════════════════════════════

@login_required
@module_permission_required('payroll', 'MANAGE')
def payroll_requests(request):
    """View payroll generation requests/history."""
    date.today()
    
    # Get all payroll periods - with error handling
    periods_by_year = {}
    total_periods = 0
    released_count = 0
    draft_count = 0
    generated_count = 0
    
    try:
        all_periods = list(PayrollPeriod.objects.all().values_list('id', 'year', 'month', 'status'))
        total_periods = len(all_periods)
        
        # Count by status
        released_count = sum(1 for p in all_periods if p[3] == 'RELEASED')
        draft_count = sum(1 for p in all_periods if p[3] == 'DRAFT')
        generated_count = sum(1 for p in all_periods if p[3] == 'GENERATED')
        
        # Now fetch full period objects for display
        for period_id in [p[0] for p in all_periods]:
            try:
                period = PayrollPeriod.objects.get(id=period_id)
                year = period.year
                if year not in periods_by_year:
                    periods_by_year[year] = []
                periods_by_year[year].append(period)
            except Exception:
                continue
    except Exception:
        periods_by_year = {}
        total_periods = 0
        released_count = 0
        draft_count = 0
        generated_count = 0
    
    context = {
        'tab': 'requests',
        'periods_by_year': periods_by_year,
        'total_periods': total_periods,
        'released_count': released_count,
        'draft_count': draft_count,
        'generated_count': generated_count,
    }
    
    return render(request, 'payroll/requests.html', context)


# ═══════════════════════════════════════════════════════════════════
#                    VET DETAILS EDITOR
# ═══════════════════════════════════════════════════════════════════

@login_required
@module_permission_required('payroll', 'MANAGE')
def vet_detail(request, staff_id):
    """
    Specialized Vet Details editor for payroll-specific data.
    Manage base salary, performance bonuses, and deductions.
    """
    staff = get_object_or_404(StaffMember, id=staff_id, is_active=True)
    today = date.today()
    
    # Import here to avoid circular import
    from .models import PayrollAuditLog
    
    if request.method == 'POST':
        old_salary = staff.salary
        
        # Update salary
        new_salary = safe_decimal(request.POST.get('base_salary', 0))
        staff.salary = new_salary
        staff.save()
        
        # Log the change
        if old_salary != new_salary:
            PayrollAuditLog.log(
                user=request.user,
                action_type=PayrollAuditLog.ActionType.VET_SALARY_UPDATED,
                description=f"Updated salary for {staff.full_name} from ₱{old_salary:,.2f} to ₱{new_salary:,.2f}",
                staff_member=staff,
                metadata={
                    'old_salary': str(old_salary),
                    'new_salary': str(new_salary)
                },
                ip_address=get_client_ip(request)
            )
        
        messages.success(request, f'Payroll data for {staff.full_name} updated successfully.')
        return redirect('payroll:vet_detail', staff_id=staff.id)
    
    # Get payroll history for this employee - with error handling
    payslip_history = []
    try:
        ps_ids = list(Payslip.objects.filter(
            employee=staff
        ).values_list('id', flat=True).order_by('-payroll_period__year', '-payroll_period__month')[:12])
        
        for ps_id in ps_ids:
            try:
                ps = Payslip.objects.select_related('payroll_period').get(id=ps_id)
                payslip_history.append(ps)
            except Exception:
                continue
    except Exception:
        payslip_history = []
    
    # Calculate YTD stats - with error handling
    ytd_stats = {
        'total_gross': Decimal('0'),
        'total_net': Decimal('0'),
        'total_deductions': Decimal('0'),
        'count': 0
    }
    try:
        stats = Payslip.objects.filter(
            employee=staff,
            payroll_period__year=today.year,
            status='RELEASED'
        ).aggregate(
            total_gross=Sum('gross_pay'),
            total_net=Sum('net_pay'),
            total_deductions=Sum('total_deductions'),
            count=Count('id')
        )
        if stats and stats['total_gross'] is not None:
            ytd_stats = stats
    except Exception:
        # Keep default values on error
        pass
    
    # Get recent audit logs for this staff - with error handling
    recent_logs = []
    try:
        log_ids = list(PayrollAuditLog.objects.filter(
            staff_member=staff
        ).values_list('id', flat=True).order_by('-created_at')[:10])
        
        for log_id in log_ids:
            try:
                log = PayrollAuditLog.objects.get(id=log_id)
                recent_logs.append(log)
            except Exception:
                continue
    except Exception:
        recent_logs = []
    
    context = {
        'tab': 'vets',
        'staff': staff,
        'payslip_history': payslip_history,
        'ytd_stats': ytd_stats,
        'recent_logs': recent_logs,
    }
    
    return render(request, 'payroll/vet_detail.html', context)


# ═══════════════════════════════════════════════════════════════════
#                    COMPREHENSIVE AUDIT LOG VIEW
# ═══════════════════════════════════════════════════════════════════

@login_required
@module_permission_required('payroll', 'MANAGE')
def audit_log_view(request):
    """
    Comprehensive audit log view showing all system actions.
    Filterable by action type, user, and date range.
    """
    from django.core.paginator import Paginator
    from .models import PayrollAuditLog
    
    # Get log IDs first to avoid decimal field access during query
    base_logs = PayrollAuditLog.objects.all().order_by('-created_at')
    
    # Apply filters to IDs only
    action_filter = request.GET.get('action_type')
    if action_filter:
        base_logs = base_logs.filter(action_type=action_filter)
    
    user_filter = request.GET.get('user')
    if user_filter:
        base_logs = base_logs.filter(user_id=user_filter)
    
    date_from = request.GET.get('date_from')
    if date_from:
        base_logs = base_logs.filter(created_at__date__gte=date_from)
    
    date_to = request.GET.get('date_to')
    if date_to:
        base_logs = base_logs.filter(created_at__date__lte=date_to)
    
    # Get filtered log IDs
    try:
        log_ids = list(base_logs.values_list('id', flat=True))
    except Exception:
        log_ids = []
    
    # Fetch individual logs with error handling
    logs = []
    for log_id in log_ids:
        try:
            log = PayrollAuditLog.objects.select_related(
                'user', 'payroll_period', 'payslip', 'staff_member'
            ).get(id=log_id)
            logs.append(log)
        except Exception:
            continue
    
    # Stats for cards - with error handling
    stats = {
        'total': 0,
        'edits': 0,
        'releases': 0,
        'salary_changes': 0,
    }
    
    try:
        stats['total'] = PayrollAuditLog.objects.all().count()
        stats['edits'] = PayrollAuditLog.objects.filter(action_type='PAYSLIP_EDITED').count()
        stats['releases'] = PayrollAuditLog.objects.filter(action_type__in=['PAYSLIP_RELEASED', 'BULK_RELEASE']).count()
        stats['salary_changes'] = PayrollAuditLog.objects.filter(action_type='VET_SALARY_UPDATED').count()
    except Exception:
        stats = {'total': 0, 'edits': 0, 'releases': 0, 'salary_changes': 0}
    
    # Pagination
    paginator = Paginator(logs, 25)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    
    # Get users for filter dropdown
    users = []
    try:
        users = list(User.objects.filter(payroll_audit_logs__isnull=False).distinct())
    except Exception:
        users = []
    
    context = {
        'tab': 'logs',
        'logs': page_obj,
        'stats': stats,
        'users': users,
    }
    
    return render(request, 'payroll/audit_log.html', context)


@login_required
@module_permission_required('payroll', 'MANAGE')
def audit_log_detail(request, log_id):
    """Display detailed information about a specific audit log entry."""
    from .models import PayrollAuditLog
    
    try:
        log = PayrollAuditLog.objects.select_related(
            'user', 'payroll_period', 'payslip', 'staff_member', 'payslip__employee'
        ).get(id=log_id)
    except PayrollAuditLog.DoesNotExist:
        from django.http import Http404
        raise Http404("Audit log not found")
    except Exception:
        from django.http import Http404
        raise Http404("Error loading audit log")
    
    context = {
        'tab': 'logs',
        'log': log,
    }
    
    return render(request, 'payroll/audit_log_detail.html', context)


# ═══════════════════════════════════════════════════════════════════
#                    HELPER FUNCTIONS
# ═══════════════════════════════════════════════════════════════════

def get_client_ip(request):
    """Extract client IP address from request."""
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        ip = x_forwarded_for.split(',')[0]
    else:
        ip = request.META.get('REMOTE_ADDR')
    return ip


# ═══════════════════════════════════════════════════════════════════
#               EXPORT FEATURES - CSV & EXCEL
# ═══════════════════════════════════════════════════════════════════

@login_required
@module_permission_required('payroll', 'MANAGE')
def export_payslips_csv(request, period_id):
    """Export payslips to CSV format."""
    import csv
    from django.http import HttpResponse
    
    period = get_object_or_404(PayrollPeriod, id=period_id)
    
    # Create CSV response
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="payslips_{period.period_display}.csv"'
    
    writer = csv.writer(response)
    
    # Write header
    writer.writerow([
        'Employee Name', 'Position', 'Branch', 'Base Salary',
        'Overtime Pay', 'Holiday Pay', 'Bonus', 'Allowance',
        'Gross Pay', 'SSS', 'PhilHealth', 'PAG-IBIG', 'Tax',
        'Cash Advance', 'Other Deductions', 'Net Pay', 'Status'
    ])
    
    # Get payslips safely - ID first approach
    payslips = []
    try:
        payslip_ids = list(period.payslips.values_list('id', flat=True))
        for ps_id in payslip_ids:
            try:
                ps = Payslip.objects.select_related('employee', 'employee__branch').get(id=ps_id)
                payslips.append(ps)
            except Exception:
                continue
    except Exception:
        payslips = []
    
    # Write data rows with error handling for decimal fields
    totals_gross = Decimal('0')
    totals_net = Decimal('0')
    
    for payslip in payslips:
        try:
            # Get decimal values with defaults
            base_salary = payslip.base_salary or Decimal('0')
            overtime_pay = payslip.overtime_pay or Decimal('0')
            holiday_pay = payslip.holiday_pay or Decimal('0')
            bonus = payslip.bonus or Decimal('0')
            allowance = payslip.allowance or Decimal('0')
            gross_pay = payslip.gross_pay or Decimal('0')
            sss = payslip.sss or Decimal('0')
            philhealth = payslip.philhealth or Decimal('0')
            pagibig = payslip.pagibig or Decimal('0')
            tax = payslip.tax or Decimal('0')
            cash_advance = payslip.cash_advance or Decimal('0')
            other_deductions = payslip.other_deductions or Decimal('0')
            net_pay = payslip.net_pay or Decimal('0')
            
            totals_gross += gross_pay
            totals_net += net_pay
            
            writer.writerow([
                payslip.employee.full_name,
                payslip.employee.position or '',
                payslip.employee.branch.name if payslip.employee.branch else '',
                f"{base_salary:.2f}",
                f"{overtime_pay:.2f}",
                f"{holiday_pay:.2f}",
                f"{bonus:.2f}",
                f"{allowance:.2f}",
                f"{gross_pay:.2f}",
                f"{sss:.2f}",
                f"{philhealth:.2f}",
                f"{pagibig:.2f}",
                f"{tax:.2f}",
                f"{cash_advance:.2f}",
                f"{other_deductions:.2f}",
                f"{net_pay:.2f}",
                payslip.get_status_display()
            ])
        except Exception:
            # Skip payslips with errors
            continue
    
    # Add summary row
    writer.writerow([])  # Blank row
    writer.writerow(['TOTALS', '', '', '', '', '', '', '',
                     f"{totals_gross:.2f}", '', '', '', '',
                     '', '', f"{totals_net:.2f}", ''])
    
    # Log the export
    try:
        from .models import PayrollAuditLog
        PayrollAuditLog.log(
            user=request.user,
            action_type=PayrollAuditLog.ActionType.SYSTEM_EXPORT,
            description=f"Exported {len(payslips)} payslips to CSV for {period.period_display}",
            payroll_period=period,
            metadata={'format': 'CSV', 'employee_count': len(payslips)}
        )
    except Exception:
        pass
    
    messages.success(request, f'Exported {len(payslips)} payslips to CSV.')
    
    return response


@login_required
@module_permission_required('payroll', 'MANAGE')
def export_payslips_excel(request, period_id):
    """Export payslips to Excel format (.xlsx)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from django.http import HttpResponse
    
    period = get_object_or_404(PayrollPeriod, id=period_id)
    
    # Get payslips safely - ID first approach
    payslips = []
    try:
        payslip_ids = list(period.payslips.values_list('id', flat=True))
        for ps_id in payslip_ids:
            try:
                ps = Payslip.objects.select_related('employee', 'employee__branch').get(id=ps_id)
                payslips.append(ps)
            except Exception:
                continue
    except Exception:
        payslips = []
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = 'Payslips'
    
    # Define styles
    header_fill = PatternFill(start_color='009688', end_color='009688', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    total_fill = PatternFill(start_color='E0F2F1', end_color='E0F2F1', fill_type='solid')
    total_font = Font(bold=True, size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Add title
    ws['A1'] = f"Payroll Report - {period.period_display}"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:Q1')
    
    # Add headers
    headers = ['Employee Name', 'Position', 'Branch', 'Base Salary',
               'Overtime Pay', 'Holiday Pay', 'Bonus', 'Allowance',
               'Gross Pay', 'SSS', 'PhilHealth', 'PAG-IBIG', 'Tax',
               'Cash Advance', 'Other Ded.', 'Net Pay', 'Status']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Set column widths
    column_widths = [20, 15, 15, 12, 12, 12, 10, 12, 12, 10, 12, 10, 10, 12, 12, 12, 10]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + col)].width = width
    
    # Add data rows with error handling
    row = 4
    totals_gross = Decimal('0')
    totals_net = Decimal('0')
    
    for payslip in payslips:
        try:
            # Get decimal values with defaults
            base_salary = payslip.base_salary or Decimal('0')
            overtime_pay = payslip.overtime_pay or Decimal('0')
            holiday_pay = payslip.holiday_pay or Decimal('0')
            bonus = payslip.bonus or Decimal('0')
            allowance = payslip.allowance or Decimal('0')
            gross_pay = payslip.gross_pay or Decimal('0')
            sss = payslip.sss or Decimal('0')
            philhealth = payslip.philhealth or Decimal('0')
            pagibig = payslip.pagibig or Decimal('0')
            tax = payslip.tax or Decimal('0')
            cash_advance = payslip.cash_advance or Decimal('0')
            other_deductions = payslip.other_deductions or Decimal('0')
            net_pay = payslip.net_pay or Decimal('0')
            
            totals_gross += gross_pay
            totals_net += net_pay
            
            ws.cell(row=row, column=1, value=payslip.employee.full_name).border = border
            ws.cell(row=row, column=2, value=payslip.employee.position or '').border = border
            ws.cell(row=row, column=3, value=payslip.employee.branch.name if payslip.employee.branch else '').border = border
            ws.cell(row=row, column=4, value=float(base_salary)).border = border
            ws.cell(row=row, column=5, value=float(overtime_pay)).border = border
            ws.cell(row=row, column=6, value=float(holiday_pay)).border = border
            ws.cell(row=row, column=7, value=float(bonus)).border = border
            ws.cell(row=row, column=8, value=float(allowance)).border = border
            ws.cell(row=row, column=9, value=float(gross_pay)).border = border
            ws.cell(row=row, column=10, value=float(sss)).border = border
            ws.cell(row=row, column=11, value=float(philhealth)).border = border
            ws.cell(row=row, column=12, value=float(pagibig)).border = border
            ws.cell(row=row, column=13, value=float(tax)).border = border
            ws.cell(row=row, column=14, value=float(cash_advance)).border = border
            ws.cell(row=row, column=15, value=float(other_deductions)).border = border
            ws.cell(row=row, column=16, value=float(net_pay)).border = border
            ws.cell(row=row, column=17, value=payslip.get_status_display()).border = border
            row += 1
        except Exception:
            # Skip payslips with errors
            continue
    
    # Add totals row
    total_row = row + 1
    ws.cell(row=total_row, column=1, value='TOTALS').font = total_font
    ws.cell(row=total_row, column=1).fill = total_fill
    ws.cell(row=total_row, column=9, value=float(totals_gross)).font = total_font
    ws.cell(row=total_row, column=9).fill = total_fill
    ws.cell(row=total_row, column=16, value=float(totals_net)).font = total_font
    ws.cell(row=total_row, column=16).fill = total_fill
    
    # Format as currency
    for row_cells in ws.iter_rows(min_row=4, max_row=total_row, min_col=4, max_col=16):
        for cell in row_cells:
            if cell.value and isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0.00'
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="payslips_{period.period_display}.xlsx"'
    
    wb.save(response)
    
    # Log the export
    try:
        from .models import PayrollAuditLog
        PayrollAuditLog.log(
            user=request.user,
            action_type=PayrollAuditLog.ActionType.SYSTEM_EXPORT,
            description=f"Exported {len(payslips)} payslips to Excel for {period.period_display}",
            payroll_period=period,
            metadata={'format': 'XLSX', 'employee_count': len(payslips)}
        )
    except Exception:
        pass
    
    messages.success(request, f'Exported {len(payslips)} payslips to Excel.')
    
    return response


# ═══════════════════════════════════════════════════════════════════
#               EMAIL PAYSLIPS FEATURE
# ═══════════════════════════════════════════════════════════════════

@login_required
@module_permission_required('payroll', 'MANAGE')
def send_payslip_email(request, payslip_id):
    """Send payslip via email to employee."""
    from django.core.mail import send_mail
    from .models import PayslipEmailLog
    
    payslip = get_object_or_404(
        Payslip.objects.select_related('employee', 'payroll_period'),
        id=payslip_id
    )
    
    # Check if employee has email
    if not payslip.employee.email:
        messages.error(request, f'{payslip.employee.full_name} does not have an email address.')
        return redirect('payroll:payslips', period_id=payslip.payroll_period.id)
    
    try:
        # Get decimal values with defaults to avoid InvalidOperation errors
        base_salary = payslip.base_salary or Decimal('0')
        overtime_pay = payslip.overtime_pay or Decimal('0')
        holiday_pay = payslip.holiday_pay or Decimal('0')
        bonus = payslip.bonus or Decimal('0')
        allowance = payslip.allowance or Decimal('0')
        gross_pay = payslip.gross_pay or Decimal('0')
        sss = payslip.sss or Decimal('0')
        philhealth = payslip.philhealth or Decimal('0')
        pagibig = payslip.pagibig or Decimal('0')
        tax = payslip.tax or Decimal('0')
        cash_advance = payslip.cash_advance or Decimal('0')
        other_deductions = payslip.other_deductions or Decimal('0')
        net_pay = payslip.net_pay or Decimal('0')
        days_worked = payslip.days_worked or 0
        days_absent = payslip.days_absent or 0
        
        # Build email content
        subject = f'Payslip for {payslip.payroll_period.period_display} - FMH Animal Clinic'
        
        message = f"""
Dear {payslip.employee.full_name},

Please find your payslip for {payslip.payroll_period.period_display} below:

═══════════════════════════════════════════════════════════════
PAYSLIP SUMMARY
═══════════════════════════════════════════════════════════════

Base Salary:          ₱{base_salary:>12,.2f}
Overtime Pay:         ₱{overtime_pay:>12,.2f}
Holiday Pay:          ₱{holiday_pay:>12,.2f}
Bonus:                ₱{bonus:>12,.2f}
Allowance:            ₱{allowance:>12,.2f}
───────────────────────────────────────────────────────────────
Gross Pay:            ₱{gross_pay:>12,.2f}

DEDUCTIONS:
SSS:                  ₱{sss:>12,.2f}
PhilHealth:           ₱{philhealth:>12,.2f}
PAG-IBIG:             ₱{pagibig:>12,.2f}
Tax (BIR):            ₱{tax:>12,.2f}
Cash Advance:         ₱{cash_advance:>12,.2f}
Other Deductions:     ₱{other_deductions:>12,.2f}
───────────────────────────────────────────────────────────────
Net Pay (Take-home):  ₱{net_pay:>12,.2f}
═══════════════════════════════════════════════════════════════

Days Worked: {days_worked} days
Days Absent: {days_absent} days

This is an automated payroll notification. Please contact the Human Resources
department if you have any questions regarding your payslip.

Best regards,
FMH Animal Clinic - Finance Department
        """
        
        # Send email
        send_mail(
            subject,
            message,
            'noreply@fmhanimalclinic.com',  # From email
            [payslip.employee.email],  # To email
            fail_silently=False,
        )
        
        # Log the email
        email_log = PayslipEmailLog.objects.create(
            payslip=payslip,
            email=payslip.employee.email,
            status=PayslipEmailLog.Status.SENT,
            sent_by=request.user,
            sent_at=timezone.now()
        )
        
        # Log the action
        from .models import PayrollAuditLog
        PayrollAuditLog.log(
            user=request.user,
            action_type=PayrollAuditLog.ActionType.SYSTEM_EXPORT,
            description=f"Sent payslip email to {payslip.employee.full_name} ({payslip.employee.email})",
            payslip=payslip,
            payroll_period=payslip.payroll_period,
            staff_member=payslip.employee,
            metadata={'email': payslip.employee.email}
        )
        
        messages.success(request, f'Payslip sent to {payslip.employee.email}')
        
    except Exception as e:
        # Log error
        try:
            email_log = PayslipEmailLog.objects.create(
                payslip=payslip,
                email=payslip.employee.email,
                status=PayslipEmailLog.Status.FAILED,
                sent_by=request.user,
                error_message=str(e)
            )
        except Exception:
            pass
        
        messages.error(request, f'Failed to send email: {str(e)}')
    
    return redirect('payroll:payslips', period_id=payslip.payroll_period.id)


@login_required
@module_permission_required('payroll', 'MANAGE')
def send_payslips_bulk(request, period_id):
    """Send payslips to all employees in a period via email."""
    from django.core.mail import send_mail
    from .models import PayslipEmailLog
    
    period = get_object_or_404(PayrollPeriod, id=period_id)
    
    # Get payslips safely - ID first approach
    payslips = []
    try:
        payslip_ids = list(period.payslips.filter(
            employee__email__isnull=False
        ).exclude(employee__email='').values_list('id', flat=True))
        
        for ps_id in payslip_ids:
            try:
                ps = Payslip.objects.select_related('employee').get(id=ps_id)
                payslips.append(ps)
            except Exception:
                continue
    except Exception:
        payslips = []
    
    sent_count = 0
    failed_count = 0
    
    for payslip in payslips:
        try:
            # Get decimal values with defaults
            gross_pay = payslip.gross_pay or Decimal('0')
            net_pay = payslip.net_pay or Decimal('0')
            
            subject = f'Payslip for {period.period_display} - FMH Animal Clinic'
            
            message = f"""
Dear {payslip.employee.full_name},

Your payslip for {period.period_display} has been generated and is ready for review.

Gross Pay: ₱{gross_pay:,.2f}
Net Pay: ₱{net_pay:,.2f}

Please contact the Finance department if you have any questions.

Best regards,
FMH Animal Clinic
            """
            
            send_mail(
                subject,
                message,
                'noreply@fmhanimalclinic.com',
                [payslip.employee.email],
                fail_silently=False,
            )
            
            # Log successful send
            try:
                PayslipEmailLog.objects.create(
                    payslip=payslip,
                    email=payslip.employee.email,
                    status=PayslipEmailLog.Status.SENT,
                    sent_by=request.user,
                    sent_at=timezone.now()
                )
            except Exception:
                pass
            
            sent_count += 1
            
        except Exception as e:
            # Log failed send
            try:
                PayslipEmailLog.objects.create(
                    payslip=payslip,
                    email=payslip.employee.email,
                    status=PayslipEmailLog.Status.FAILED,
                    sent_by=request.user,
                    error_message=str(e)
                )
            except Exception:
                pass
            failed_count += 1
    
    # Log the bulk send action
    try:
        from .models import PayrollAuditLog
        PayrollAuditLog.log(
            user=request.user,
            action_type=PayrollAuditLog.ActionType.SYSTEM_EXPORT,
            description=f"Sent {sent_count} payslip emails for {period.period_display} ({failed_count} failed)",
            payroll_period=period,
            metadata={'sent': sent_count, 'failed': failed_count}
        )
    except Exception:
        pass
    
    if sent_count > 0:
        messages.success(request, f'Sent {sent_count} payslip emails.')
    if failed_count > 0:
        messages.warning(request, f'{failed_count} emails failed to send.')

    return redirect('payroll:payslips', period_id=period.id)


# ═══════════════════════════════════════════════════════════════════
#                    STAFF PAYSLIP VIEWS (Self-Service)
# ═══════════════════════════════════════════════════════════════════

@login_required
@special_permission_required('can_view_own_payslips')
def my_payslips_view(request):
    """
    List all released payslips for the currently logged-in staff member.
    Only shows payslips from RELEASED payroll periods.
    """
    # Get the staff profile linked to the current user
    try:
        staff_profile = request.user.staff_profile
    except AttributeError:
        messages.error(request, 'You do not have a staff profile associated with your account.')
        return redirect('admin_dashboard')

    if staff_profile is None:
        messages.error(request, 'Your account is not linked to a staff member record.')
        return redirect('admin_dashboard')

    # Only show payslips from RELEASED payroll periods
    payslips = Payslip.objects.filter(
        employee=staff_profile,
        payroll_period__status=PayrollPeriod.Status.RELEASED
    ).select_related(
        'payroll_period'
    ).order_by('-payroll_period__year', '-payroll_period__month')

    # Calculate summary stats
    total_earned = payslips.aggregate(total=Sum('net_pay'))['total'] or 0

    context = {
        'payslips': payslips,
        'total_payslips': payslips.count(),
        'total_earned': total_earned,
        'staff_member': staff_profile,
    }

    return render(request, 'payroll/my_payslips.html', context)


@login_required
@special_permission_required('can_view_own_payslips')
def my_payslip_detail_view(request, pk):
    """
    View a single payslip detail - user can ONLY view their own payslips.
    Additional security: only released payslips are viewable.
    """
    # Get the staff profile linked to the current user
    try:
        staff_profile = request.user.staff_profile
    except AttributeError:
        raise Http404("Payslip not found")

    if staff_profile is None:
        raise Http404("Payslip not found")

    # Fetch payslip - must belong to this user
    payslip = get_object_or_404(Payslip, pk=pk, employee=staff_profile)

    # Security check: only released payslips can be viewed
    if payslip.payroll_period.status != PayrollPeriod.Status.RELEASED:
        raise Http404("Payslip not found")

    context = {
        'payslip': payslip,
        'is_own_payslip': True,  # Flag for template to hide edit controls
    }

    return render(request, 'payroll/my_payslip_detail.html', context)
